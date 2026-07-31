# -*- coding: utf-8 -*-
"""Reporte Financiero del proyecto.

Una fila por (proyecto, vigencia) con los valores agregados de toda la cadena:
CDP -> compromiso -> obligacion -> reserva.

La reserva se asocia a la vigencia de su CDP de origen (cdp_origen), NO a su
vigencia de constitucion: una reserva constituida en v+1 arrastra la ejecucion
pendiente de v, asi que se reporta bajo v (la vigencia del CDP/RP que la origino).

Cada etapa se agrega por separado (una consulta por tabla, agrupada por
proyecto+vigencia) y se cruzan en Python por la llave: nunca se anida un Sum
sobre otro join, para no inflar ni colapsar valores.
"""
from collections import defaultdict
from decimal import Decimal

from django.db.models import DurationField, ExpressionWrapper, F, Max, Min, Sum

from .models import (CdpImputacion, CompromisoImputacion, ContratoImputacion,
                     ObligacionImputacion, Proyecto, ReservaImputacion)

CERO = Decimal("0")

# (clave, encabezado, tipo, grupo)   tipo: texto | fecha | money | entero
# El grupo alimenta la fila 1 del Excel (celdas combinadas por bloque).
G_BASICOS = "Datos básicos"
G_DISP = "Disponibilidad presupuestal"
G_REG = "Registro presupuestal"
G_OBLI = "Obligaciones y pagos"
G_RES = "Reservas"
G_CONTR = "Contratación"
COLUMNAS = [
    ("bpin",                 "BPIN",                            "texto",  G_BASICOS),
    ("nombre",               "Nombre del proyecto",             "texto",  G_BASICOS),
    ("clasificaciones",      "Clasificacion",                   "texto",  G_BASICOS),
    ("dependencia_responsable", "Dependencia responsable",      "texto",  G_BASICOS),
    ("vigencia",             "Vigencia",                        "texto",  G_BASICOS),
    ("fecha_primer_cdp",     "FechaPrimerCDP",                  "fecha",  G_DISP),
    ("valor_certificado",    "ValorCertificado",                "money",  G_DISP),
    ("valor_disp_def",       "ValorDisponibilidadDefinitiva",   "money",  G_DISP),
    ("saldo_certf",          "ValorSaldoCertificado",           "money",  G_DISP),
    ("fecha_primer_rp",      "FechaPrimerRP",                   "fecha",  G_REG),
    ("valor_registro",       "ValorRegistro",                   "money",  G_REG),
    ("valor_compromiso_def", "ValorCompromisoDefinitivo",       "money",  G_REG),
    ("saldo_rp",             "ValorSaldoCompromiso",            "money",  G_REG),
    ("fecha_primera_obli",   "FechaPrimeraObligacion",          "fecha",  G_OBLI),
    ("valor_obligacion_def", "ValorObligacionDefinitiva",       "money",  G_OBLI),
    ("saldo_obli",           "ValorSaldoObligacion",            "money",  G_OBLI),
    ("pagos",                "ValorPagos",                      "money",  G_OBLI),
    ("fecha_reserva",        "FechaReserva",                    "fecha",  G_RES),
    ("valor_reserva",        "ValorReserva",                    "money",  G_RES),
    ("valor_reserva_def",    "ValorReservaDefinitiva",          "money",  G_RES),
    ("obligaciones_reserva", "ObligacionesReservasDefinitivas", "money",  G_RES),
    ("saldo_reserva",        "ValorSaldoReservas",              "money",  G_RES),
    ("pagos_reserva",        "ValorPagoReservas",               "money",  G_RES),
    ("concepto",             "Tipo de orden de gasto",          "texto",  G_CONTR),
    ("fecha_firma_primer_contrato", "FechaFirmaPrimerContrato", "fecha",  G_CONTR),
    ("fecha_inicio_contrato", "FechaInicioContrato",            "fecha",  G_CONTR),
    ("duracion_dias",        "DuracionDias (contrato mas largo)", "entero", G_CONTR),
    ("contratos",            "Contratos (cantidad)",            "entero", G_CONTR),
]


def _indexar(qs):
    """{(proyecto_id, vigencia): fila} a partir de un values(...).annotate(...)."""
    return {(r["proyecto"], r["v"]): r for r in qs}


def construir(vigencias=None, bpines=None, nombre=None):
    """Devuelve la lista de filas (dict) ordenadas por BPIN y vigencia.

    Filtros opcionales: vigencia (aplicada a la etapa de cada tabla), BPIN
    (lista exacta) y nombre del proyecto (contiene).
    """
    def base(qs, campo_vigencia):
        qs = qs.filter(proyecto__isnull=False)
        if vigencias:
            qs = qs.filter(**{f"{campo_vigencia}__in": vigencias})
        if bpines:
            qs = qs.filter(proyecto__bpin__in=bpines)
        if nombre:
            qs = qs.filter(proyecto__nombre__icontains=nombre)
        return qs

    cdp = _indexar(
        base(CdpImputacion.objects, "cdp__vigencia")
        .values("proyecto", v=F("cdp__vigencia"))
        .annotate(fecha_primer_cdp=Min("cdp__fecha_disp"),
                  valor_certificado=Sum("valor_certificado"),
                  valor_disp_def=Sum("valor_disponibilidad_def"),
                  saldo_certf=Sum("saldo_certf")))
    comp = _indexar(
        base(CompromisoImputacion.objects, "compromiso__vigencia")
        .values("proyecto", v=F("compromiso__vigencia"))
        .annotate(fecha_primer_rp=Min("compromiso__fecha_reg"),
                  valor_registro=Sum("valor_registro"),
                  valor_compromiso_def=Sum("valor_compromiso_def"),
                  saldo_rp=Sum("saldo_rp")))
    obli = _indexar(
        base(ObligacionImputacion.objects, "obligacion__vigencia")
        .values("proyecto", v=F("obligacion__vigencia"))
        .annotate(fecha_primera_obli=Min("obligacion__fecha_obli"),
                  valor_obligacion_def=Sum("valor_obligacion"),
                  saldo_obli=Sum("saldo_obli"),
                  pagos=Sum("pagos")))
    # La reserva va bajo la vigencia de su CDP de origen (v-1 respecto a su constitucion)
    res = _indexar(
        base(ReservaImputacion.objects, "cdp_origen__vigencia")
        .values("proyecto", v=F("cdp_origen__vigencia"))
        .annotate(fecha_reserva=Min("reserva__fecha_reserva"),
                  valor_reserva=Sum("valor_reserva"),
                  valor_reserva_def=Sum("valor_reserva_def"),
                  obligaciones_reserva=Sum("obligaciones_reserva"),
                  saldo_reserva=Sum("saldo_reserva"),
                  pagos_reserva=Sum("pagos_reserva")))

    # Datos de contrato por (proyecto, vigencia), via ContratoImputacion.vigencia.
    # Duracion del proyecto = la del contrato mas largo (Max de fecha_final - fecha_inicio).
    # Min/Max no se inflan aunque un contrato tenga varias imputaciones.
    contr = _indexar(
        base(ContratoImputacion.objects, "vigencia")
        .values("proyecto", v=F("vigencia"))
        .annotate(fecha_firma_primer_contrato=Min("contrato__fecha_firma"),
                  fecha_inicio_contrato=Min("contrato__fecha_inicio"),
                  duracion=Max(ExpressionWrapper(
                      F("contrato__fecha_final") - F("contrato__fecha_inicio"),
                      output_field=DurationField()))))

    # Texto agregado (distinct, separado por coma)
    conceptos = defaultdict(set)
    for r in (base(ObligacionImputacion.objects, "obligacion__vigencia")
              .values("proyecto", v=F("obligacion__vigencia"),
                      val=F("obligacion__tipo_orden_gasto__nombre")).distinct()):
        if r["val"]:
            conceptos[(r["proyecto"], r["v"])].add(r["val"])
    contratos = defaultdict(set)
    for r in (base(ContratoImputacion.objects, "vigencia")
              .values("proyecto", v=F("vigencia"), val=F("contrato__nro_contrato")).distinct()):
        if r["val"]:
            contratos[(r["proyecto"], r["v"])].add(r["val"])

    llaves = set(cdp) | set(comp) | set(obli) | set(res)

    # Metadatos del proyecto (nombre y clasificaciones), una sola pasada
    ids = {pid for pid, _ in llaves}
    meta = {}
    for p in (Proyecto.objects.filter(id__in=ids)
              .select_related("dependencia_responsable").prefetch_related("clasificaciones")):
        meta[p.id] = (p.bpin or "", p.nombre or "",
                      ", ".join(c.nombre for c in p.clasificaciones.all()),
                      p.dependencia_responsable.nombre if p.dependencia_responsable_id else "")

    fuentes = (
        (cdp,  ("fecha_primer_cdp", "valor_certificado", "valor_disp_def", "saldo_certf")),
        (comp, ("fecha_primer_rp", "valor_registro", "valor_compromiso_def", "saldo_rp")),
        (obli, ("fecha_primera_obli", "valor_obligacion_def", "saldo_obli", "pagos")),
        (res,  ("fecha_reserva", "valor_reserva", "valor_reserva_def",
                "obligaciones_reserva", "saldo_reserva", "pagos_reserva")),
    )
    tipos = {c[0]: c[2] for c in COLUMNAS}

    filas = []
    for pid, v in sorted(llaves, key=lambda k: (meta.get(k[0], ("",))[0], k[1] or 0)):
        bpin, nombre, clasif, dep_resp = meta.get(pid, ("", "", "", ""))
        c = contr.get((pid, v), {})
        dur = c.get("duracion")
        fila = {"bpin": bpin, "nombre": nombre, "clasificaciones": clasif,
                "dependencia_responsable": dep_resp, "vigencia": v,
                "concepto": ", ".join(sorted(conceptos.get((pid, v), ()))),
                "fecha_firma_primer_contrato": c.get("fecha_firma_primer_contrato"),
                "fecha_inicio_contrato": c.get("fecha_inicio_contrato"),
                "duracion_dias": dur.days if hasattr(dur, "days") else dur,
                "contratos": len(contratos.get((pid, v), ()))}
        for indice, campos in fuentes:
            reg = indice.get((pid, v), {})
            for c in campos:
                valor = reg.get(c)
                # money vacio -> 0; fecha vacia -> None
                fila[c] = valor if valor is not None else (CERO if tipos[c] == "money" else None)
        filas.append(fila)
    return filas


def a_excel(filas):
    from itertools import groupby
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte financiero"

    fill_grupo = PatternFill("solid", fgColor="12501A")   # combinado (grupos)
    fill_col = PatternFill("solid", fgColor="196B24")      # encabezados de columna
    blanco = Font(color="FFFFFF", bold=True, size=11)
    lado = Side(style="thin", color="000000")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)   # bordes negros
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fila 1: grupos combinados. El fondo/borde va a TODAS las celdas del grupo
    # ANTES de combinar (una celda combinada no acepta estilos en las tapadas).
    inicio = 1
    for grupo, cols in groupby(COLUMNAS, key=lambda c: c[3]):
        fin = inicio + sum(1 for _ in cols) - 1
        for col in range(inicio, fin + 1):
            cel = ws.cell(row=1, column=col)
            cel.fill = fill_grupo
            cel.border = borde
            cel.alignment = centro
        celda = ws.cell(row=1, column=inicio, value=grupo)
        celda.font = blanco
        if fin > inicio:
            ws.merge_cells(start_row=1, start_column=inicio, end_row=1, end_column=fin)
        inicio = fin + 1

    # Fila 2: encabezados de columna
    for col, (clave, encabezado, tipo, grupo) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=2, column=col, value=encabezado)
        c.fill = fill_col
        c.font = blanco
        c.alignment = centro
        c.border = borde

    # Datos desde la fila 3. Todo centrado verticalmente (al medio); el texto
    # largo (nombre, clasificacion, conceptos) se ajusta para que se lea bien.
    ali_num = Alignment(horizontal="right", vertical="center")
    ali_fecha = Alignment(horizontal="center", vertical="center")
    ali_texto = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for i, fila in enumerate(filas, start=3):
        for col, (clave, encabezado, tipo, grupo) in enumerate(COLUMNAS, start=1):
            c = ws.cell(row=i, column=col, value=fila.get(clave))
            c.border = borde
            if tipo in ("money", "entero"):
                c.number_format = "#,##0"
                c.alignment = ali_num
            elif tipo == "fecha":
                c.number_format = "DD/MM/YYYY"
                c.alignment = ali_fecha
            else:
                c.alignment = ali_texto

    # Formato de Tabla de Excel (filtros y estructura) sobre encabezado + datos.
    # Estilo sin nombre para NO pisar los colores propios de la cabecera.
    ultima_fila = len(filas) + 2
    ref = f"A2:{get_column_letter(len(COLUMNAS))}{ultima_fila}"
    tabla = Table(displayName="ReporteFinanciero", ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(name=None, showRowStripes=False,
                                          showColumnStripes=False)
    ws.add_table(tabla)

    anchos = {"texto": 34, "fecha": 15, "money": 18, "entero": 12}
    anchos_col = {"nombre": 55}   # el nombre del proyecto va mas ancho
    for col, (clave, encabezado, tipo, grupo) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = anchos_col.get(clave, anchos[tipo])
    # Sin panes congelados (a pedido)
    return wb
