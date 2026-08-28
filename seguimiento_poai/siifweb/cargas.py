# -*- coding: utf-8 -*-
"""Carga de los reportes xlsx de SIIFWEB a la base de datos.

El alcance es el SEGUIMIENTO DE PROYECTOS: solo se cargan los reportes cuyas filas
llevan PROYECTO. Quedan fuera ingresos, la bitacora y las cuentas por pagar.

Reglas de normalizacion (ver NOTAS_CRUCES.md del proyecto de analisis):
  - Identificadores como texto, sin sufijo ".0"
  - Valores con coma decimal -> Decimal
  - Fechas como datetime del xlsx o texto dd/mm/aa
  - Las reservas apuntan a CDPs de la vigencia ANTERIOR
  - Obligacion cuyo RP no existe en su vigencia se busca en v-1 (ejecucion de reserva)
"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import load_workbook

from .models import (BpinProceso, Cdp, CdpImputacion, CentroCosto, Clasificacion, Compromiso,
                     CompromisoImputacion, Contrato, ContratoSecop, DependenciaResponsable,
                     ContratoActa, ContratoImputacion, Fuente, Obligacion, ObligacionImputacion,
                     OrdenGasto, ProcesoSecop, Proyecto, Reserva, ReservaImputacion, Rubro, Tercero)

CERO = Decimal("0")


# ---------- normalizacion ----------

def texto(valor):
    """Identificador o texto limpio. Los numeros pierden el sufijo .0"""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip().removesuffix(".0")


def decimal(valor):
    """Valor monetario. Acepta float, int y texto con coma decimal."""
    if valor is None or valor == "":
        return CERO
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    try:
        return Decimal(str(valor).strip().replace(".", "").replace(",", ".")
                       if str(valor).count(",") == 1 and str(valor).count(".") > 1
                       else str(valor).strip().replace(",", "."))
    except InvalidOperation:
        return CERO


def fecha(valor):
    """Fecha del xlsx (datetime) o texto dd/mm/aa | dd/mm/aaaa."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    for formato in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(valor).strip(), formato).date()
        except ValueError:
            continue
    return None


def stream(archivo):
    """Contenido del FieldFile como stream en memoria.

    Sirve para almacenamiento local y remoto (S3): FileField.path NO existe en
    S3 (lanza NotImplementedError). openpyxl lee igual de un file-object.
    """
    with archivo.open("rb") as fh:
        return BytesIO(fh.read())


def leer_filas(ruta, hoja=None, tabla=None, columnas=None):
    """Devuelve la lista de filas del xlsx como diccionarios.

    hoja: nombre preferido; si no existe se usa la primera del libro.
    tabla: nombre de una TABLA de Excel. Se busca en todas las hojas del libro y se
        leen solo sus celdas, asi que el archivo puede llamarse como sea y la tabla
        estar donde sea: lo que manda es el nombre de la tabla. Es lo que hace falta
        para un libro de trabajo con decenas de hojas, donde caer en la primera
        cargaria cualquier cosa en silencio.
    columnas: encabezados que el reporte debe traer; si falta alguno se aborta.
    """
    if tabla:
        # ws.tables no existe en modo read_only: hay que abrir el libro completo
        libro = load_workbook(ruta, data_only=True)
        rango = next(((h, h.tables[tabla]) for h in libro.worksheets if tabla in h.tables), None)
        if rango is None:
            libro.close()
            raise ValueError(f"El archivo no contiene la tabla '{tabla}'. "
                             f"Hojas del libro: {', '.join(libro.sheetnames)}")
        # ws.tables[nombre] devuelve el objeto Table (con .ref) o el rango en texto,
        # segun la version de openpyxl
        celdas = rango[0][getattr(rango[1], "ref", rango[1])]
        encabezado = [str(c.value).strip() if c.value is not None else "" for c in celdas[0]]
        filas = [dict(zip(encabezado, [c.value for c in fila])) for fila in celdas[1:]
                 if any(c.value is not None for c in fila)]
    else:
        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = libro[hoja] if hoja and hoja in libro.sheetnames else libro.worksheets[0]
        iterador = hoja.iter_rows(values_only=True)
        encabezado = [str(c).strip() if c is not None else "" for c in next(iterador)]
        filas = [dict(zip(encabezado, fila)) for fila in iterador
                 if any(c is not None for c in fila)]
    libro.close()

    if columnas:
        faltan = [c for c in columnas if c not in encabezado]
        if faltan:
            raise ValueError(f"Al reporte le faltan columnas: {', '.join(faltan)}")
    return filas


# ---------- catalogos con cache ----------

VACIOS = {"", "0", "SIN PROYECTO", "NO", "NONE", "NAN"}


def util(valor):
    """Un nombre sirve si no esta vacio ni es un marcador de ausencia."""
    limpio = texto(valor)
    return limpio if limpio.upper() not in VACIOS else ""


class Catalogos:
    """get_or_create con cache, y ENRIQUECIMIENTO entre fuentes.

    Un mismo rubro, fuente, proyecto o tercero aparece en varios reportes, y no todos
    traen la misma calidad de datos: el BPIN llega sin nombre desde los consolidados y
    con nombre desde el historial; el NIT llega sin razon social desde la bitacora y con
    ella desde reservas. Cuando un reporte trae un dato que falta, se completa.
    """

    def __init__(self):
        self._rubros = {}
        self._fuentes = {}
        self._centros = {}
        self._proyectos = {}
        self._terceros = {}
        self._ordenes = {}
        self.enriquecidos = 0
        # Visibilidad para el equipo: que proyectos toco esta carga
        self.enganchados_a_manuales = set()   # BPIN que alguien creo a mano y ahora tiene ejecucion
        self.proyectos_nuevos = set()         # BPIN que aparecio por primera vez en el reporte

    def resumen_proyectos(self):
        partes = []
        if self.enganchados_a_manuales:
            partes.append(f"{len(self.enganchados_a_manuales)} proyectos del equipo con ejecucion")
        if self.proyectos_nuevos:
            partes.append(f"{len(self.proyectos_nuevos)} BPIN nuevos del reporte")
        return " | ".join(partes)

    def _completar(self, objeto, campo, valor):
        """Rellena un campo vacio del catalogo con el dato que trae este reporte."""
        nuevo = util(valor)
        if nuevo and not getattr(objeto, campo):
            setattr(objeto, campo, nuevo[:255])
            objeto.save(update_fields=[campo])
            self.enriquecidos += 1

    def rubro(self, codigo, nombre=None, tipo=None):
        codigo = texto(codigo)
        if tipo is None:  # 1.x = ingreso, 2.x = gasto
            tipo = "I" if codigo.startswith("1") else "G"
        if codigo not in self._rubros:
            self._rubros[codigo], _ = Rubro.objects.get_or_create(
                codigo=codigo, defaults={"nombre": util(nombre) or codigo, "tipo": tipo})
        objeto = self._rubros[codigo]
        if objeto.nombre == objeto.codigo:  # se creo sin nombre real
            self._completar(objeto, "nombre", nombre)
        return objeto

    def fuente(self, codigo, nombre=None):
        codigo = texto(codigo)
        if codigo not in self._fuentes:
            self._fuentes[codigo], _ = Fuente.objects.get_or_create(
                codigo=codigo[:5], defaults={"nombre": util(nombre) or codigo})
        objeto = self._fuentes[codigo]
        if objeto.nombre == objeto.codigo:
            self._completar(objeto, "nombre", nombre)
        return objeto

    def centro(self, codigo, nombre=None):
        codigo = texto(codigo) or "0"
        if codigo not in self._centros:
            self._centros[codigo], _ = CentroCosto.objects.get_or_create(
                codigo=codigo[:5], defaults={"nombre": util(nombre) or codigo})
        objeto = self._centros[codigo]
        if objeto.nombre == objeto.codigo:
            self._completar(objeto, "nombre", nombre)
        return objeto

    def proyecto(self, bpin, nombre=None, dependencia=None):
        """'0' o vacio = funcionamiento -> None (ausencia de relacion).

        Busca por BPIN: si el equipo ya creo el proyecto a mano, la carga se engancha
        a ese registro y respeta lo que hayan escrito. Solo crea uno nuevo si el BPIN
        no existia, y lo marca como detectado en el reporte.
        """
        bpin = texto(bpin)
        if bpin in ("", "0"):
            return None
        if bpin not in self._proyectos:
            objeto, nuevo = Proyecto.objects.get_or_create(
                bpin=bpin, defaults={"origen": Proyecto.Origen.SIIFWEB})
            self._proyectos[bpin] = objeto
            if not nuevo and objeto.origen == Proyecto.Origen.MANUAL:
                self.enganchados_a_manuales.add(bpin)
            elif nuevo:
                self.proyectos_nuevos.add(bpin)
        objeto = self._proyectos[bpin]
        self._completar(objeto, "nombre", nombre)
        if dependencia is not None and objeto.dependencia_id is None:
            objeto.dependencia = dependencia
            objeto.save(update_fields=["dependencia"])
            self.enriquecidos += 1
        return objeto

    def tercero(self, nit, nombre=None):
        nit = texto(nit) or "0"
        if nit not in self._terceros:
            self._terceros[nit], _ = Tercero.objects.get_or_create(
                codigo=nit[:30], defaults={"nombre": util(nombre)})
        objeto = self._terceros[nit]
        self._completar(objeto, "nombre", nombre)
        return objeto

    def orden_gasto(self, nombre):
        nombre = texto(nombre) or "SIN TIPO"
        if nombre not in self._ordenes:
            self._ordenes[nombre], _ = OrdenGasto.objects.get_or_create(nombre=nombre)
        return self._ordenes[nombre]


# ---------- reemplazo por vigencia ----------

def reemplazar(modelo_imputacion, filtro):
    """Borra las imputaciones de la vigencia antes de volver a cargarla.

    La vigencia en curso se descarga una y otra vez, y cada descarga es la version
    buena. Sin esto las imputaciones se acumularian y los valores se duplicarian:
    las cabeceras estan protegidas por su UniqueConstraint, pero las lineas no.

    Solo se borran las LINEAS. Las cabeceras se actualizan en sitio (upsert) porque
    otros documentos apuntan a ellas: borrar un Cdp rompería los RP que lo citan.
    """
    return modelo_imputacion.objects.filter(**filtro).delete()[0]


def upsert(modelo, objetos, llave, campos):
    """Crea los documentos nuevos y actualiza los que ya existian."""
    modelo.objects.bulk_create(list(objetos), update_conflicts=True,
                               unique_fields=llave, update_fields=campos, batch_size=1000)


def armar_detalle(base, previas, cat=None, extra=""):
    """Mensaje del resultado: que se cargo, que se reemplazo y a que proyectos toco."""
    partes = [base]
    if extra:
        partes.append(extra)
    if previas:
        partes.append(f"reemplazo {previas} imputaciones de la carga anterior")
    if cat:
        resumen = cat.resumen_proyectos()
        if resumen:
            partes.append(resumen)
        if cat.enriquecidos:
            partes.append(f"{cat.enriquecidos} catalogos enriquecidos")
    return " | ".join(partes)


# ---------- cargadores ----------

@transaction.atomic
def cargar_cdp(carga):
    filas = leer_filas(stream(carga.archivo))
    cat = Catalogos()
    vigencia = carga.vigencia

    documentos = {}
    for f in filas:
        nro = texto(f["NRO_CDP"])
        if nro and nro not in documentos:
            documentos[nro] = Cdp(
                vigencia=vigencia, nro_cdp=nro,
                fecha_disp=fecha(f["FECHA_DISP"]),
                objeto_cert=texto(f.get("OBJETO_CERT")),
                centro_costo=cat.centro(f.get("ID_CENTROCOSTO"), f.get("SOLICITANTE")),
            )
    previas = reemplazar(CdpImputacion, {"cdp__vigencia": vigencia})
    upsert(Cdp, documentos.values(), ["vigencia", "nro_cdp"],
           ["fecha_disp", "objeto_cert", "centro_costo"])
    mapa = {c.nro_cdp: c for c in Cdp.objects.filter(vigencia=vigencia)}

    imputaciones = [
        CdpImputacion(
            reporte=carga, cdp=mapa[texto(f["NRO_CDP"])],
            rubro=cat.rubro(f["IDENTIFICACION_PRESUPUESTAL"], f.get("NOMBRE_RUBRO")),
            fuente=cat.fuente(f["FONDO"], f.get("NOMBRE_FONDO")),
            proyecto=cat.proyecto(f.get("PROYECTO")),
            valor_certificado=decimal(f["VALOR_CERTIFICADO"]),
            valor_disponibilidad_def=decimal(f["VALOR_DISPONIBILIDAD_DEF"]),
            saldo_certf=decimal(f["SALDO_CERTF"]),
        ) for f in filas if texto(f["NRO_CDP"]) in mapa
    ]
    CdpImputacion.objects.bulk_create(imputaciones, batch_size=1000)
    return len(filas), armar_detalle(f"{len(documentos)} CDPs, {len(imputaciones)} imputaciones",
                                     previas, cat)


@transaction.atomic
def cargar_compromisos(carga):
    filas = leer_filas(stream(carga.archivo))
    cat = Catalogos()
    vigencia = carga.vigencia
    cdps = {c.nro_cdp: c for c in Cdp.objects.filter(vigencia=vigencia)}

    documentos = {}
    for f in filas:
        nro = texto(f["NRO_RP"])
        if nro and nro not in documentos:
            documentos[nro] = Compromiso(
                vigencia=vigencia, nro_rp=nro,
                fecha_reg=fecha(f["FECHA_REG"]),
                acto_admon=texto(f.get("ACTO_ADMON")),
                objeto_reg=texto(f.get("OBJETO_REG")),
                centro_costo=cat.centro(f.get("CCOSTO"), f.get("SOLICITANTE")),
            )
    previas = reemplazar(CompromisoImputacion, {"compromiso__vigencia": vigencia})
    upsert(Compromiso, documentos.values(), ["vigencia", "nro_rp"],
           ["fecha_reg", "acto_admon", "objeto_reg", "centro_costo"])
    mapa = {c.nro_rp: c for c in Compromiso.objects.filter(vigencia=vigencia)}

    imputaciones, sin_cdp = [], 0
    for f in filas:
        nro_cdp = texto(f.get("NRODOC_CDP"))
        if nro_cdp not in cdps:
            sin_cdp += 1
            continue
        imputaciones.append(CompromisoImputacion(
            reporte=carga, compromiso=mapa[texto(f["NRO_RP"])], cdp=cdps[nro_cdp],
            rubro=cat.rubro(f["IDENTIFICACION_PRESUPUESTAL"], f.get("NOMBRE_RUBRO")),
            fuente=cat.fuente(f["FONDO"], f.get("NOMBRE_FONDO")),
            proyecto=cat.proyecto(f.get("PROYECTO")),
            valor_registro=decimal(f["VALOR_REGISTRO"]),
            valor_compromiso_def=decimal(f["VALOR_COMPROMISO_DEF"]),
            saldo_rp=decimal(f["SALDO_RP"]),
        ))
    CompromisoImputacion.objects.bulk_create(imputaciones, batch_size=1000)
    aviso = f"{sin_cdp} filas sin CDP en la vigencia, omitidas" if sin_cdp else ""
    return len(filas), armar_detalle(f"{len(documentos)} RPs, {len(imputaciones)} imputaciones",
                                     previas, cat, aviso)


@transaction.atomic
def cargar_obligaciones(carga):
    filas = leer_filas(stream(carga.archivo))
    cat = Catalogos()
    vigencia = carga.vigencia
    rps = {c.nro_rp: c for c in Compromiso.objects.filter(vigencia=vigencia)}
    rps_anterior = {c.nro_rp: c for c in Compromiso.objects.filter(vigencia=vigencia - 1)}

    documentos = {}
    for f in filas:
        nro = texto(f["NRO_OBLIGACION"])
        if nro and nro not in documentos:
            documentos[nro] = Obligacion(
                vigencia=vigencia, nro_obligacion=nro,
                fecha_obli=fecha(f["FECHA_OBLI"]),
                objeto_oblig=texto(f.get("OBJETO_OBLIG")),
                # El consolidado trae NIT y razon social: de paso enriquece el catalogo
                beneficiario=(cat.tercero(f.get("NIT"), f.get("BENEFICIARIO"))
                              if texto(f.get("NIT")) else None),
                tipo_orden_gasto=cat.orden_gasto(f.get("TIPO_ORDEN_GASTO")),
                nro_orden_gasto=texto(f.get("NRO_ORDEN_GASTO")) or None,
                prefijo_orden=texto(f.get("PREFIJO_ORDEN")) or None,
                orden_pago=texto(f.get("ORDEN_PAGO")) or None,
                centro_costo=cat.centro(f.get("CCOSTO"), f.get("NOMBRE_CCOSTO")),
            )
    previas = reemplazar(ObligacionImputacion, {"obligacion__vigencia": vigencia})
    upsert(Obligacion, documentos.values(), ["vigencia", "nro_obligacion"],
           ["fecha_obli", "objeto_oblig", "beneficiario", "tipo_orden_gasto",
            "nro_orden_gasto", "prefijo_orden", "orden_pago", "centro_costo"])
    mapa = {c.nro_obligacion: c for c in Obligacion.objects.filter(vigencia=vigencia)}

    imputaciones, de_reserva, huerfanas = [], 0, 0
    for f in filas:
        nro_rp = texto(f.get("NRO_RP"))
        rp = rps.get(nro_rp)
        if rp is None:
            # Regla verificada: RP ausente en su vigencia y presente en v-1 = ejecucion de reserva
            rp = rps_anterior.get(nro_rp)
            if rp is None:
                huerfanas += 1
                continue
            de_reserva += 1
        imputaciones.append(ObligacionImputacion(
            reporte=carga, obligacion=mapa[texto(f["NRO_OBLIGACION"])], compromiso=rp,
            rubro=cat.rubro(f["IDENTIFICACION_PRESUPUESTAL"], f.get("NOMBRE_RUBRO")),
            fuente=cat.fuente(f["FONDO"], f.get("NOMBRE_FONDO")),
            proyecto=cat.proyecto(f.get("PROYECTO")),
            valor_obligacion=decimal(f["VALOR_OBLI_DEF"]),
            saldo_obli=decimal(f["SALDO_OBLI"]),
            pagos=decimal(f["PAGOS"]),
        ))
    ObligacionImputacion.objects.bulk_create(imputaciones, batch_size=1000)
    extra = []
    if de_reserva:
        extra.append(f"{de_reserva} contra RP de {vigencia - 1} (ejecucion de reserva)")
    if huerfanas:
        extra.append(f"{huerfanas} sin RP en ninguna vigencia (omitidas)")
    return len(filas), armar_detalle(
        f"{len(documentos)} obligaciones, {len(imputaciones)} imputaciones",
        previas, cat, ", ".join(extra))


@transaction.atomic
def cargar_reservas(carga):
    filas = leer_filas(stream(carga.archivo))
    cat = Catalogos()
    vigencia = carga.vigencia
    # Salto temporal: la reserva de v se constituye con CDPs de v-1
    cdps = {c.nro_cdp: c for c in Cdp.objects.filter(vigencia=vigencia - 1)}

    documentos = {}
    for f in filas:
        nro = texto(f["NRO_RESERVA"])
        if nro and nro not in documentos:
            documentos[nro] = Reserva(
                vigencia=vigencia, nro_reserva=nro,
                fecha_reserva=fecha(f["FECHA_RESERVA"]),
                beneficiario=cat.tercero(f.get("NIT"), f.get("BENEFICIARIO")),
                objeto_reserva=texto(f.get("OBJETO_RESERVA")),
                acto_admon=texto(f.get("ACTO_ADMON")),
            )
    previas = reemplazar(ReservaImputacion, {"reserva__vigencia": vigencia})
    upsert(Reserva, documentos.values(), ["vigencia", "nro_reserva"],
           ["fecha_reserva", "beneficiario", "objeto_reserva", "acto_admon"])
    mapa = {c.nro_reserva: c for c in Reserva.objects.filter(vigencia=vigencia)}

    imputaciones, sin_cdp = [], 0
    for f in filas:
        cdp = cdps.get(texto(f.get("NUMERO_CDP")))
        if cdp is None:
            sin_cdp += 1
        imputaciones.append(ReservaImputacion(
            reporte=carga, reserva=mapa[texto(f["NRO_RESERVA"])], cdp_origen=cdp,
            rubro=cat.rubro(f["IDENTIFICACION_PRESUPUESTAL"], f.get("NOMBRE_RUBRO")),
            fuente=cat.fuente(f["FONDO"], f.get("NOMBRE_FONDO")),
            proyecto=cat.proyecto(f.get("PROYECTO")),
            valor_reserva=decimal(f["VALOR_RESERVA"]),
            valor_reserva_def=decimal(f["VALOR_RESERVA_DEF"]),
            obligaciones_reserva=decimal(f["OBLIGACIONES_RESERVA"]),
            pagos_reserva=decimal(f["PAGOS_RESERVA"]),
            saldo_reserva=decimal(f["SALDO_RESERVA"]),
        ))
    ReservaImputacion.objects.bulk_create(imputaciones, batch_size=1000)
    aviso = f"{sin_cdp} sin CDP de {vigencia - 1} cargado" if sin_cdp else ""
    return len(filas), armar_detalle(
        f"{len(documentos)} reservas, {len(imputaciones)} imputaciones", previas, cat, aviso)

@transaction.atomic
def cargar_historial(carga):
    """Historial de orden de gasto 2: una sabana que se parte en tres tablas.

    A diferencia de los consolidados, este reporte se descarga por RANGO completo
    (el contrato cruza vigencias) y cada fila del archivo es una de tres cosas:
      - solo ficha del contrato
      - un evento de pago (acta)      -> VALOR_PAGO diligenciado
      - una imputacion presupuestal   -> NRO_COMPROBANTEPPTAL diligenciado
    """
    filas = leer_filas(stream(carga.archivo))
    cat = Catalogos()

    # Rango completo: cada descarga sustituye a la anterior. Se borran primero los
    # hijos (PROTECT) y luego las fichas.
    previas = ContratoImputacion.objects.all().delete()[0]
    previas += ContratoActa.objects.all().delete()[0]
    Contrato.objects.all().delete()

    def llave(f):
        return (texto(f.get("TIPO_CONTRATO")), texto(f.get("NRO_CONTRATO")),
                fecha(f.get("FECHA_FIRMA")), texto(f.get("TERCERO")))

    # 1. Fichas: una por contrato
    fichas = {}
    for f in filas:
        k = llave(f)
        if k[1] and k[2] and k not in fichas:
            fichas[k] = Contrato(
                reporte=carga,
                tipo_contrato=k[0][:250], nro_contrato=k[1][:250], fecha_firma=k[2],
                tercero=cat.tercero(f.get("TERCERO"), f.get("CONTRATISTA")),
                interventor=texto(f.get("INTERVENTOR"))[:250] or None,
                descripcion=texto(f.get("DESCRIPCION")) or None,
                valor_contrato=decimal(f.get("VALOR_CONTRATO")) or None,
                fecha_inicio=fecha(f.get("FECHA_INICIO")),
                fecha_final=fecha(f.get("FECHA_FINAL1")),
                dependencia=cat.centro(f.get("DIVISION2"), f.get("NOMBRE")),
            )
    Contrato.objects.bulk_create(list(fichas.values()), batch_size=1000)
    mapa = {(c.tipo_contrato, c.nro_contrato, c.fecha_firma, c.tercero.codigo): c
            for c in Contrato.objects.select_related("tercero").all()}

    # 2. Actas: las filas con pago
    actas = []
    for f in filas:
        valor = decimal(f.get("VALOR_PAGO"))
        contrato = mapa.get(llave(f))
        if contrato is None or valor == CERO:
            continue
        actas.append(ContratoActa(
            reporte=carga, contrato=contrato,
            nro_orden=texto(f.get("NRO_ORDEN"))[:20],
            tipo_orden=texto(f.get("TIPO_ORDEN"))[:250],
            concepto=texto(f.get("CONCEPTO")),
            fecha_pago=fecha(f.get("FECHA_PAGO")),
            nrodoc_acta=texto(f.get("NRODOC_ACTA"))[:20],
            valor_pago=valor,
        ))
    ContratoActa.objects.bulk_create(actas, batch_size=1000)

    # 3. Imputaciones: las filas con comprobante presupuestal.
    #    NRODOC es el numero del RP y PREFIJO su vigencia.
    rps = {(c.vigencia, c.nro_rp): c for c in Compromiso.objects.all()}
    imputaciones, con_rp = [], 0
    for f in filas:
        comprobante = texto(f.get("NRO_COMPROBANTEPPTAL"))
        contrato = mapa.get(llave(f))
        if contrato is None or not comprobante:
            continue
        vigencia = int(texto(f.get("PREFIJO"))) if texto(f.get("PREFIJO")).isdigit() else None
        rp = rps.get((vigencia, texto(f.get("NRODOC"))))
        if rp is not None:
            con_rp += 1
        imputaciones.append(ContratoImputacion(
            reporte=carga, vigencia=vigencia, contrato=contrato, compromiso=rp,
            nro_comprobante=comprobante[:20],
            rubro=cat.rubro(f["RUBRO"], f.get("NOMBRERUBRO") or f.get("RUBRO")),
            fuente=cat.fuente(f["RECURSO"], f.get("NOMBRE_RECURSO")),
            # Este reporte es la unica fuente del NOMBRE del proyecto y de su dependencia:
            # los consolidados solo traen el BPIN
            proyecto=cat.proyecto(f.get("PROYECTO"), f.get("NOMBRE_PROYECTO"),
                                  cat.centro(f.get("ID_CENTROCOSTO"), f.get("NOMBRE_CC"))),
        ))
    ContratoImputacion.objects.bulk_create(imputaciones, batch_size=1000)

    return len(filas), armar_detalle(
        f"{len(fichas)} contratos, {len(actas)} actas, {len(imputaciones)} imputaciones",
        previas, cat, f"{con_rp} enganchadas a un RP cargado")


@transaction.atomic
def cargar_poai(carga):
    """Cruce POAI: enriquece el catalogo de proyectos. Carga OPCIONAL.

    El reporte trae una fila por indicador y vigencia, asi que cada BPIN se repite
    (194 proyectos en 380 filas). Los atributos del proyecto son identicos en todas
    sus filas -verificado-, de modo que se deduplica quedandose con la primera.

    Reglas de escritura:
      - nombre: solo si SIIFWEB no lo trajo (el historial es la fuente primaria)
      - dependencia responsable y clasificaciones: SIIFWEB no los tiene, asi que este
        reporte es la fuente autoritativa y siempre los actualiza
    """
    filas = leer_filas(stream(carga.archivo), hoja="Completo")

    unicos = {}
    for f in filas:
        bpin = texto(f.get("BPIN"))
        if bpin and bpin not in unicos:
            unicos[bpin] = f

    cache_dep, cache_clas = {}, {}

    def dependencia(valor):
        """NO se parte por coma: estos nombres las llevan dentro
        ('Mujer, Juventud, e Inclusion Social')."""
        nombre = util(valor)
        if not nombre:
            return None
        if nombre not in cache_dep:
            cache_dep[nombre], _ = DependenciaResponsable.objects.get_or_create(nombre=nombre[:150])
        return cache_dep[nombre]

    def clasificaciones(valor):
        """Aqui si: en una celda pueden venir varias separadas por coma
        ('POAI 2026, Recursos del Balance')."""
        salida = []
        for parte in str(util(valor)).split(","):
            nombre = parte.strip()
            if not nombre:
                continue
            if nombre not in cache_clas:
                cache_clas[nombre], _ = Clasificacion.objects.get_or_create(nombre=nombre[:120])
            salida.append(cache_clas[nombre])
        return salida

    creados = nombres = otros = manuales = 0
    for bpin, f in unicos.items():
        proyecto, nuevo = Proyecto.objects.get_or_create(
            bpin=bpin, defaults={"origen": Proyecto.Origen.POAI})
        creados += nuevo
        if not nuevo and proyecto.origen == Proyecto.Origen.MANUAL:
            manuales += 1
        cambios = []

        nombre = util(f.get("NombreProyecto"))
        if nombre and not proyecto.nombre:  # no pisa lo que ya trajo SIIFWEB
            proyecto.nombre = nombre
            cambios.append("nombre")
            nombres += 1

        dep = dependencia(f.get("Dependencia Responsable"))
        if dep and proyecto.dependencia_responsable_id != dep.pk:
            proyecto.dependencia_responsable = dep
            cambios.append("dependencia_responsable")
            otros += 1

        if cambios:
            proyecto.save(update_fields=cambios)

        clases = clasificaciones(f.get("Clasificación"))
        if clases:
            proyecto.clasificaciones.set(clases)
            otros += len(clases)

    detalle = (f"{len(unicos)} proyectos en el reporte ({len(filas)} filas antes de deduplicar) | "
               f"{nombres} nombres completados, {otros} campos POAI escritos | "
               f"{len(cache_dep)} dependencias responsables, {len(cache_clas)} clasificaciones")
    if creados:
        detalle += f" | {creados} proyectos nuevos (estan en el POAI, sin ejecucion en SIIFWEB)"
    if manuales:
        detalle += f" | {manuales} proyectos del equipo completados"
    return len(filas), detalle


# ---------- SECOP II ----------

# Las 23 columnas de la tabla consolidada. Si el libro no las trae todas, la carga
# se aborta antes de escribir nada.
COLUMNAS_SECOP = (
    "BPIN", "Año", "ID Proceso", "ID Contrato", "ID Portafolio", "Validacion BPIN",
    "Proceso de Compra", "Referencia del Contrato", "Estado Contrato",
    "Descripcion del Proceso", "Fecha de Firma", "Fecha de Inicio del Contrato",
    "Fecha de Fin del Contrato", "Documento Proveedor", "Valor del Contrato",
    "URLProceso", "Objeto del Contrato", "Fecha de Publicacion del Proceso",
    "Fecha de Ultima Publicación", "Fecha de Recepcion de Respuestas",
    "Fecha de Apertura de Respuesta", "Fecha de Apertura Efectiva",
    "Estado del Procedimiento",
)

SIN_CONTRATO = "No Definido"   # marcador del proceso que aun no adjudica


@transaction.atomic
def cargar_secop(carga):
    """Consolidado 'BPIN por proceso' de SECOP II: tres bases del DNP en una tabla.

    La tabla se busca por NOMBRE (`BPIN_por_proceso`), no por hoja ni por archivo:
    sirve cualquier libro que la contenga con sus 23 columnas.

    Cada fila es un BPIN con su proceso y, si ya adjudico, su contrato. De ahi salen
    las tres tablas: el proceso (fechas de publicacion y estado del procedimiento),
    el contrato electronico (referencia, proveedor, valor, fechas) y la fila que los
    ata al proyecto por BPIN.

    Se descarga por rango completo, como el historial de contratos: cada carga
    reemplaza a la anterior.
    """
    filas = leer_filas(stream(carga.archivo), tabla="BPIN_por_proceso", columnas=COLUMNAS_SECOP)
    cat = Catalogos()

    previas = BpinProceso.objects.all().delete()[0]
    ContratoSecop.objects.all().delete()
    ProcesoSecop.objects.all().delete()

    # 1. Procesos. 'Proceso de Compra' es la misma columna que 'ID Portafolio' en el
    #    lado de los contratos: se verifica y no se guarda dos veces.
    procesos, discrepancias = {}, 0
    for f in filas:
        pid = texto(f.get("ID Proceso"))
        compra = texto(f.get("Proceso de Compra"))
        portafolio = texto(f.get("ID Portafolio"))
        if compra and compra != portafolio:
            discrepancias += 1
        if not pid or pid in procesos:
            continue
        procesos[pid] = ProcesoSecop(
            reporte=carga,
            id_proceso=pid[:60],
            id_portafolio=portafolio[:60],
            estado_procedimiento=texto(f.get("Estado del Procedimiento"))[:60],
            fecha_publicacion=fecha(f.get("Fecha de Publicacion del Proceso")),
            fecha_ultima_publicacion=fecha(f.get("Fecha de Ultima Publicación")),
            fecha_recepcion_respuestas=fecha(f.get("Fecha de Recepcion de Respuestas")),
            fecha_apertura_respuestas=fecha(f.get("Fecha de Apertura de Respuesta")),
            fecha_apertura_efectiva=fecha(f.get("Fecha de Apertura Efectiva")),
        )
    ProcesoSecop.objects.bulk_create(list(procesos.values()), batch_size=1000)
    mapa_procesos = {p.id_proceso: p for p in ProcesoSecop.objects.all()}

    # 2. Contratos electronicos. El mismo contrato se repite en tantas filas como
    #    BPIN financie, con los mismos datos -verificado: 0 inconsistencias-, asi
    #    que se toma la primera aparicion.
    contratos = {}
    for f in filas:
        cid = texto(f.get("ID Contrato"))
        proceso = mapa_procesos.get(texto(f.get("ID Proceso")))
        if not cid or cid == SIN_CONTRATO or cid in contratos or proceso is None:
            continue
        nit = texto(f.get("Documento Proveedor"))
        contratos[cid] = ContratoSecop(
            reporte=carga,
            id_contrato=cid[:60],
            proceso=proceso,
            referencia=texto(f.get("Referencia del Contrato"))[:120],
            estado=texto(f.get("Estado Contrato"))[:60],
            objeto=texto(f.get("Objeto del Contrato")),
            descripcion_proceso=texto(f.get("Descripcion del Proceso")),
            # El consolidado trae el NIT pero no la razon social: el nombre lo
            # completa cualquier otro reporte que traiga ese tercero.
            proveedor=cat.tercero(nit) if nit else None,
            valor=decimal(f.get("Valor del Contrato")),
            fecha_firma=fecha(f.get("Fecha de Firma")),
            fecha_inicio=fecha(f.get("Fecha de Inicio del Contrato")),
            fecha_fin=fecha(f.get("Fecha de Fin del Contrato")),
            url_proceso=texto(f.get("URLProceso")),
        )
    ContratoSecop.objects.bulk_create(list(contratos.values()), batch_size=1000)
    mapa_contratos = {c.id_contrato: c for c in ContratoSecop.objects.all()}

    # 3. La fila del consolidado. No se crean proyectos desde SECOP: si el BPIN no
    #    esta en el catalogo, la fila entra con proyecto nulo y el BPIN queda a la
    #    vista para que el equipo decida.
    catalogo = {p.bpin: p for p in Proyecto.objects.exclude(bpin__isnull=True)}
    vistas, huerfanos = set(), set()
    sin_contrato = repetidas = sin_proyecto = 0
    lineas = []
    for f in filas:
        bpin = texto(f.get("BPIN"))
        proceso = mapa_procesos.get(texto(f.get("ID Proceso")))
        if not bpin or proceso is None:
            continue
        contrato = mapa_contratos.get(texto(f.get("ID Contrato")))
        llave = (bpin, proceso.id_proceso, contrato.id_contrato if contrato else None)
        if llave in vistas:       # el archivo trae filas repetidas exactas
            repetidas += 1
            continue
        vistas.add(llave)
        proyecto = catalogo.get(bpin)
        if proyecto is None:
            huerfanos.add(bpin)
            sin_proyecto += 1
        if contrato is None:
            sin_contrato += 1
        anio = texto(f.get("Año"))
        lineas.append(BpinProceso(
            reporte=carga,
            bpin=bpin[:25],
            proyecto=proyecto,
            proceso=proceso,
            contrato_secop=contrato,
            anio=int(anio) if anio.isdigit() else None,
            validacion_bpin=texto(f.get("Validacion BPIN"))[:20],
        ))
    BpinProceso.objects.bulk_create(lineas, batch_size=1000)

    bpines = {b for b, _, _ in vistas}
    detalle = armar_detalle(
        f"{len(procesos)} procesos, {len(contratos)} contratos, "
        f"{len(lineas)} filas BPIN de {len(bpines)} BPIN distintos",
        previas, cat,
        f"{sin_contrato} filas sin contrato adjudicado")
    if huerfanos:
        detalle += (f" | {sin_proyecto} filas con BPIN que no esta en el catalogo: "
                    f"{', '.join(sorted(huerfanos))}")
    if repetidas:
        detalle += f" | {repetidas} filas repetidas descartadas"
    if discrepancias:
        detalle += f" | OJO: {discrepancias} filas donde 'Proceso de Compra' difiere de 'ID Portafolio'"
    return len(filas), detalle


PROCESADORES = {
    "cdp": cargar_cdp,
    "compromisos": cargar_compromisos,
    "obligaciones": cargar_obligaciones,
    "reservas": cargar_reservas,
    "historial": cargar_historial,
    "poai": cargar_poai,
    "secop": cargar_secop,
}


def procesar(carga):
    """Ejecuta la carga y deja el resultado en el propio registro."""
    procesador = PROCESADORES.get(carga.tipo_reporte)
    if procesador is None:
        carga.estado = carga.Estado.ERROR
        carga.mensaje = f"No hay cargador para el tipo '{carga.tipo_reporte}'"
        carga.save(update_fields=["estado", "mensaje"])
        return False
    try:
        filas, detalle = procesador(carga)
        carga.filas = filas
        carga.estado = carga.Estado.PROCESADO
        carga.mensaje = detalle
    except Exception as error:  # noqa: BLE001 - el mensaje va al registro para diagnostico
        carga.estado = carga.Estado.ERROR
        carga.mensaje = f"{type(error).__name__}: {error}"
    carga.save(update_fields=["filas", "estado", "mensaje"])
    return carga.estado == carga.Estado.PROCESADO
