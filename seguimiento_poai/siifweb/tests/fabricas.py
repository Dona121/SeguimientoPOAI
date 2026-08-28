# -*- coding: utf-8 -*-
"""Piezas para armar los escenarios de las pruebas.

El libro de SECOP se construye aqui en memoria y no se lee del archivo real: las
pruebas tienen que correr en cualquier maquina, y ademas conviene poder fabricar
los casos raros (fila repetida, proceso sin contrato, BPIN fuera del catalogo) que
en el archivo de verdad aparecen mezclados.
"""
import hashlib
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from siifweb import cargas
from siifweb.models import (CargaReporte, Cdp, CdpImputacion, CentroCosto, Compromiso,
                            CompromisoImputacion, Contrato, ContratoActa, ContratoImputacion,
                            Fuente, Proyecto, Rubro, Tercero)

COLUMNAS = cargas.COLUMNAS_SECOP

# Una fila valida y completa; cada prueba cambia solo lo que le interesa
FILA = {
    "BPIN": "202500000000001",
    "Año": "2025",
    "ID Proceso": "CO1.REQ.1",
    "ID Contrato": "CO1.PCCNTR.1",
    "ID Portafolio": "CO1.BDOS.1",
    "Validacion BPIN": "Validado",
    "Proceso de Compra": "CO1.BDOS.1",
    "Referencia del Contrato": "CPS-001-2025",
    "Estado Contrato": "En ejecución",
    "Descripcion del Proceso": "PRESTACION DE SERVICIOS DE PRUEBA",
    "Fecha de Firma": date(2025, 9, 25),
    "Fecha de Inicio del Contrato": date(2025, 9, 26),
    "Fecha de Fin del Contrato": date(2025, 12, 23),
    "Documento Proveedor": "1005639763",
    "Valor del Contrato": 9000000,
    "URLProceso": "https://community.secop.gov.co/Public/Tendering/1",
    "Objeto del Contrato": "PRESTACION DE SERVICIOS DE PRUEBA",
    "Fecha de Publicacion del Proceso": date(2025, 9, 24),
    "Fecha de Ultima Publicación": date(2025, 9, 24),
    "Fecha de Recepcion de Respuestas": None,
    "Fecha de Apertura de Respuesta": None,
    "Fecha de Apertura Efectiva": None,
    "Estado del Procedimiento": "Seleccionado",
}


def fila(**cambios):
    """Una fila del consolidado con los cambios que pida la prueba."""
    nueva = dict(FILA)
    nueva.update(cambios)
    return nueva


def sin_contrato(**cambios):
    """Fila de un proceso publicado que todavia no adjudica."""
    vacias = {c: None for c in ("Referencia del Contrato", "Estado Contrato",
                                "Descripcion del Proceso", "Fecha de Firma",
                                "Fecha de Inicio del Contrato", "Fecha de Fin del Contrato",
                                "Documento Proveedor", "Valor del Contrato", "URLProceso",
                                "Objeto del Contrato", "Proceso de Compra")}
    vacias["ID Contrato"] = cargas.SIN_CONTRATO
    vacias["Estado del Procedimiento"] = "Evaluación"
    vacias.update(cambios)
    return fila(**vacias)


def libro(filas, tabla="BPIN_por_proceso", hoja="Secop II", columnas=COLUMNAS, ruido_abajo=False):
    """Un xlsx con la tabla en una hoja que NO es la primera.

    Asi se comprueba lo que pidio el usuario: que la carga encuentre la tabla por su
    nombre, en el libro que sea y en la hoja que sea.
    """
    wb = Workbook()
    senuelo = wb.active
    senuelo.title = "Hoja de otro proceso"
    senuelo.append(["PERIODO", "EFICACIA"])
    senuelo.append(["20260601", "0.023"])

    ws = wb.create_sheet(hoja)
    ws.append(list(columnas))
    for f in filas:
        ws.append([f.get(c) for c in columnas])
    if tabla:
        ref = f"A1:{get_column_letter(len(columnas))}{len(filas) + 1}"
        ws.add_table(Table(displayName=tabla, ref=ref))
    if ruido_abajo:
        # Debajo de la tabla, fuera de su rango: no debe entrar en la carga
        ws.append(["RUIDO"] + [None] * (len(columnas) - 1))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def carga_secop(filas=None, contenido=None, procesar=True, **kwargs):
    """Crea el registro de carga con su archivo y, salvo que se pida lo contrario, lo procesa."""
    datos = contenido if contenido is not None else libro(filas or [fila()], **kwargs).getvalue()
    registro = CargaReporte(tipo_reporte=CargaReporte.TipoReporte.SECOP, vigencia=None,
                            hash=hashlib.sha256(datos + str(len(datos)).encode()).hexdigest())
    registro.archivo.save("consolidado.xlsx", ContentFile(datos), save=True)
    if procesar:
        cargas.procesar(registro)
        registro.refresh_from_db()
    return registro


# ---------- catalogos y cadena de SIIFWEB ----------

def catalogos():
    rubro, _ = Rubro.objects.get_or_create(codigo="2.3.2.01.01",
                                           defaults={"nombre": "Rubro de prueba", "tipo": "G"})
    fuente, _ = Fuente.objects.get_or_create(codigo="1", defaults={"nombre": "Recursos propios"})
    centro, _ = CentroCosto.objects.get_or_create(codigo="16", defaults={"nombre": "Infraestructura"})
    return rubro, fuente, centro


def proyecto(bpin="202500000000001", nombre="Proyecto de prueba", **kwargs):
    return Proyecto.objects.create(bpin=bpin, nombre=nombre, **kwargs)


def tercero(codigo="1005639763", nombre="Contratista de prueba"):
    objeto, _ = Tercero.objects.get_or_create(codigo=codigo, defaults={"nombre": nombre})
    return objeto


def cadena(proy, vigencia=2025, valor=Decimal("10000000.00"), nro="1"):
    """CDP -> RP -> obligacion del proyecto, con una imputacion cada uno."""
    rubro, fuente, centro = catalogos()
    from siifweb.models import Obligacion, ObligacionImputacion, OrdenGasto

    cdp = Cdp.objects.create(centro_costo=centro, vigencia=vigencia, nro_cdp=nro,
                             fecha_disp=date(vigencia, 1, 15), objeto_cert="Objeto del CDP")
    CdpImputacion.objects.create(cdp=cdp, rubro=rubro, fuente=fuente, proyecto=proy,
                                 valor_certificado=valor, valor_disponibilidad_def=valor,
                                 saldo_certf=Decimal("0.00"))
    rp = Compromiso.objects.create(centro_costo=centro, vigencia=vigencia, nro_rp=nro,
                                   fecha_reg=date(vigencia, 2, 1), acto_admon="CONTRATO 1",
                                   objeto_reg="Objeto del registro")
    CompromisoImputacion.objects.create(compromiso=rp, cdp=cdp, rubro=rubro, fuente=fuente,
                                        proyecto=proy, valor_registro=valor,
                                        valor_compromiso_def=valor, saldo_rp=Decimal("0.00"))
    orden, _ = OrdenGasto.objects.get_or_create(nombre="CONTRATO")
    obligacion = Obligacion.objects.create(centro_costo=centro, tipo_orden_gasto=orden,
                                           vigencia=vigencia, nro_obligacion=nro,
                                           fecha_obli=date(vigencia, 3, 1), nro_orden_gasto=nro,
                                           objeto_oblig="Objeto de la obligacion",
                                           beneficiario=tercero())
    # Pagado a la mitad a proposito: si pagos y obligado fueran iguales, una prueba
    # que sume la columna equivocada pasaria igual
    pagado = valor / 2
    ObligacionImputacion.objects.create(obligacion=obligacion, compromiso=rp, rubro=rubro,
                                        fuente=fuente, proyecto=proy, valor_obligacion=valor,
                                        saldo_obli=valor - pagado, pagos=pagado)
    return cdp, rp, obligacion


def contrato_siifweb(proy, imputaciones=2, actas=2, valor_acta=Decimal("500000.00"),
                     valor=Decimal("10000000.00"), compromiso=None, nro="2666"):
    """Un contrato del historial con varias imputaciones y varias actas.

    Dos imputaciones del mismo proyecto mas dos actas es justo el caso que duplicaba
    el pagado cuando se sumaba con Sum() sobre dos relaciones a la vez.
    """
    rubro, fuente, centro = catalogos()
    contrato = Contrato.objects.create(
        tipo_contrato="CONTRATO DE PRESTACION", nro_contrato=nro,
        fecha_firma=date(2025, 10, 31), tercero=tercero(),
        interventor="Interventor de prueba", descripcion="Objeto del contrato de prueba",
        valor_contrato=valor, fecha_inicio=date(2025, 11, 1), fecha_final=date(2025, 12, 31),
        dependencia=centro)
    for i in range(imputaciones):
        ContratoImputacion.objects.create(vigencia=2025, contrato=contrato, compromiso=compromiso,
                                          nro_comprobante=str(i + 1), rubro=rubro, fuente=fuente,
                                          proyecto=proy)
    for i in range(actas):
        ContratoActa.objects.create(contrato=contrato, nro_orden=str(i + 1), tipo_orden="PAGO",
                                    concepto="Pago parcial", fecha_pago=date(2025, 12, i + 1),
                                    nrodoc_acta=str(i + 1), valor_pago=valor_acta)
    return contrato
